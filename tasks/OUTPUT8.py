from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os

def process_output7(aisle_id):
    """
    Process the OUTPUT7.xlsx file for the specified aisle by applying borders to all cells 
    and saving the result as OUTPUT8.

    Args:
        aisle_id (str): Aisle identifier (e.g., 'A09').

    Returns:
        str: Path to the processed OUTPUT8 file for the aisle.
    """
    try:
        # Use environment variables for folder paths
        OUTPUT_FOLDER = os.getenv("OUTPUT_FOLDER", "output")  # Default to "output"

        # File paths
        input_file = os.path.join(OUTPUT_FOLDER, f"OUTPUT7_{aisle_id}.xlsx")  # Input file for the aisle
        output_file = os.path.join(OUTPUT_FOLDER, f"OUTPUT8_{aisle_id}.xlsx")  # Output file for the aisle

        # Load the OUTPUT7 workbook
        workbook = load_workbook(input_file)
        sheet = workbook.active  # Get the active worksheet

        # Define a normal border (thin for all sides)
        normal_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Apply borders to all used cells
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = normal_border  # Apply the thin border to each cell

        # Save the updated workbook as OUTPUT8
        workbook.save(output_file)
        print(f"Processed file for {aisle_id} saved as {output_file}")

        return output_file

    except Exception as e:
        print(f"Error processing aisle {aisle_id}: {e}")
        raise Exception(f"Failed to process OUTPUT8 for aisle {aisle_id}: {e}")
