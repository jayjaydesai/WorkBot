import openpyxl
from openpyxl.styles import PatternFill, Alignment
import os

def process_bulk_file(bulk_file_path, aisle_id):
    """
    Process the BULK.xlsx file for the specified aisle and save the results to OUTPUT4.

    Args:
        bulk_file_path (str): Path to the uploaded BULK.xlsx file.
        aisle_id (str): Aisle identifier (e.g., 'A09').

    Returns:
        str: Path to the processed OUTPUT4 file for the aisle.
    """
    try:
        # Use environment variables for folder paths
        LOCATIONS_FOLDER = os.getenv("LOCATIONS_FOLDER", "locations")  # Default to "locations"
        OUTPUT_FOLDER = os.getenv("OUTPUT_FOLDER", "output")  # Default to "output"

        # File paths
        input_aisle_file = os.path.join(LOCATIONS_FOLDER, f"{aisle_id}.xlsx")  # Aisle-specific file
        output_file_path = os.path.join(OUTPUT_FOLDER, f"OUTPUT4_{aisle_id}.xlsx")  # Output file path

        # Load BULK and aisle-specific files
        bulk_wb = openpyxl.load_workbook(bulk_file_path)
        bulk_sheet = bulk_wb.active

        # Create a mapping for location and license plate, ensuring no duplicates
        location_to_license = {}
        for row in range(2, bulk_sheet.max_row + 1):  # Iterate over BULK file rows
            location = bulk_sheet[f"C{row}"].value  # Location column
            license_plate = bulk_sheet[f"E{row}"].value  # License plate column
            if location:
                # Add license plate to the set (ensuring unique values)
                if location not in location_to_license:
                    location_to_license[location] = set()
                location_to_license[location].add(license_plate)

        # Convert sets to newline-separated strings with a hyphen at the end of each line
        for location in location_to_license:
            location_to_license[location] = "\n".join(f"{plate}-" for plate in sorted(location_to_license[location]))

        # Process the aisle-specific file
        aisle_wb = openpyxl.load_workbook(input_aisle_file)
        aisle_sheet = aisle_wb.active

        # Define Light Grey fill for cells with multiple license plates
        light_grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Populate license plate data
        for row in range(2, aisle_sheet.max_row + 1, 2):  # Even rows for locations
            for col in range(3, 31):  # Columns C to AD
                location_cell = aisle_sheet.cell(row=row, column=col)
                license_cell = aisle_sheet.cell(row=row + 1, column=col)  # Just below the location cell

                if location_cell.value in location_to_license:
                    license_plate_value = location_to_license[location_cell.value]
                    license_cell.value = license_plate_value  # Assign formatted license plates
                    license_cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping

                    # Apply light grey fill if there are multiple license plates
                    if "\n" in license_plate_value:  # Check if there are multiple plates
                        license_cell.fill = light_grey_fill

        # Save the updated workbook as OUTPUT4
        aisle_wb.save(output_file_path)
        print(f"Processed file for {aisle_id} saved as {output_file_path}")

        return output_file_path

    except Exception as e:
        print(f"Error processing aisle {aisle_id}: {e}")
        raise Exception(f"Failed to process aisle {aisle_id}: {e}")
