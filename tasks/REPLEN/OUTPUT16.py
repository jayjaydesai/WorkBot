import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


def create_combined_output16(upload_folder, output_folder):
    """
    Combine the functionality of REPLENOUTPUT15 and OUTPUT16:
    - Merge BULK.xlsx, AS01011.xlsx, COVERAGE.xlsx, and optionally EXPORT.xlsx.
    - Add calculated columns for replenishment.
    - Update "Replen Status" based on "Stock" > 0.
    - Filter rows based on conditions for "Replen Status," "Stock Status," "Location," and "Ugly."
    - Add "Total Posted Qty" column grouped by "Item Number."
    - Add "Number of Item Number" column.
    - Add "Diff2," "% Required of Posted Qty," "Stock For Replen," "Decision," and "Pallet Qty-Replen Stock" columns.
    - Add "Ratio" column as the percentage of "Posted Quantity" over "Stock For Replen."
    - Save the final result as OUTPUT16.xlsx.

    Args:
        upload_folder (str): Path to the folder containing input files.
        output_folder (str): Path to save the final output file.
    """
    try:
        # Resolve paths
        upload_folder = Path(upload_folder).resolve()
        output_folder = Path(output_folder).resolve()

        print(f"Resolved upload folder: {upload_folder}")
        print(f"Resolved output folder: {output_folder}")

        # Ensure output folder exists
        output_folder.mkdir(parents=True, exist_ok=True)

        # Define required and optional files
        required_files = {
            "bulk_file": "BULK.xlsx",
            "as_file": "AS01011.xlsx",
            "coverage_file": "COVERAGE.xlsx",
        }
        optional_file = "EXPORT.xlsx"

        # Check required files
        missing_files = [
            filename for key, filename in required_files.items()
            if not (upload_folder / filename).exists()
        ]
        if missing_files:
            raise FileNotFoundError(f"Missing required files: {', '.join(missing_files)}")

        # Load BULK.xlsx
        print("Loading BULK.xlsx...")
        bulk_df = pd.read_excel(upload_folder / required_files["bulk_file"])
        bulk_df.columns = bulk_df.columns.str.lower().str.strip()

        # Load AS01011.xlsx
        print("Loading AS01011.xlsx...")
        as_df = pd.read_excel(upload_folder / required_files["as_file"], usecols=["Item number", "Available physical"])
        as_df.columns = as_df.columns.str.lower().str.strip()
        as_df["available physical"] = as_df["available physical"].fillna(0)

        # Load COVERAGE.xlsx
        print("Loading COVERAGE.xlsx...")
        coverage_df = pd.read_excel(upload_folder / required_files["coverage_file"])
        coverage_df.columns = coverage_df.columns.str.lower().str.strip()

        # Merge BULK.xlsx with COVERAGE.xlsx on "item number"
        print("Merging BULK.xlsx and COVERAGE.xlsx...")
        merged_df = pd.merge(bulk_df, coverage_df, on="item number", how="left")

        # Merge the result with AS01011.xlsx on "item number"
        print("Merging with AS01011.xlsx...")
        merged_df = pd.merge(merged_df, as_df, on="item number", how="left")

        # Load EXPORT.xlsx if available
        export_file = upload_folder / optional_file
        if export_file.exists():
            print("Loading EXPORT.xlsx...")
            export_df = pd.read_excel(export_file, usecols=["Item Number", "Stock"])
            export_df.columns = export_df.columns.str.lower().str.strip()
            print("Merging with EXPORT.xlsx...")
            merged_df = pd.merge(merged_df, export_df, on="item number", how="left")
        else:
            print(f"Warning: {optional_file} not found. Proceeding without it.")
            merged_df["stock"] = None  # Add a "stock" column with empty values if EXPORT.xlsx is missing

        # Replace blank "Available Physical" with 0
        print("Replacing blank 'Available Physical' with 0...")
        merged_df["available physical"] = merged_df["available physical"].fillna(0)

        # Add calculated columns
        print("Adding calculated columns...")
        merged_df["rmin - available physical"] = merged_df["rmin"] - merged_df["available physical"]
        merged_df["replen status"] = merged_df["rmin - available physical"].apply(
            lambda x: "Replen Required" if x > 0 else "Replen not Required"
        )
        merged_df["difference"] = merged_df["rcoverage"] - merged_df["available physical"]

        # Update "Replen Status" based on "Stock"
        print("Updating 'Replen Status' based on 'Stock' > 0...")
        merged_df["replen status"] = merged_df.apply(
            lambda row: "Replen Required" if row["stock"] > 0 else row["replen status"], axis=1
        )

        # Filter rows
        print("Filtering rows...")
        filtered_df = merged_df[
            (merged_df["replen status"] != "Replen not Required") &
            (merged_df["stock status"] != "Available") &
            (merged_df["location"] != "AS Replens") &
            ~((merged_df["ugly"] == "UGLY") & ((merged_df["stock"].isna()) | (merged_df["stock"] <= 0)))
        ]

        # Add "Total Posted Qty" column
        filtered_df["total posted qty"] = filtered_df.groupby("item number")["posted quantity"].transform("sum")

        # Add "Number of Item Number" column
        filtered_df["number of item number"] = filtered_df["item number"].map(filtered_df["item number"].value_counts())

        # Add "Diff2" and "% Required of Posted Qty" columns
        filtered_df["diff2"] = filtered_df["total posted qty"] - filtered_df["difference"]
        filtered_df["% required of posted qty"] = (
            (filtered_df["difference"] / filtered_df["total posted qty"]) * 100
        ).fillna(0).round(2).astype(str) + "%"

        # Add "Stock For Replen" column
        print("Calculating 'Stock For Replen' column...")
        filtered_df["stock for replen"] = filtered_df.apply(
            lambda row: row["total posted qty"] if row["diff2"] <= 11 else row["difference"], axis=1
        )
        filtered_df["stock for replen"] += filtered_df["stock"].fillna(0)

        # Add "Decision" column
        print("Adding 'Decision' column...")
        filtered_df["decision"] = filtered_df["number of item number"].apply(
            lambda x: "Good to Go" if x == 1 else ""
        )

        # Add "Pallet Qty-Replen Stock" column
        print("Adding 'Pallet Qty-Replen Stock' column...")
        filtered_df["pallet qty-replen stock"] = filtered_df["posted quantity"] - filtered_df["stock for replen"]

        # Add "Ratio" column
        print("Adding 'Ratio' column...")
        filtered_df["ratio"] = (filtered_df["posted quantity"] / filtered_df["stock for replen"].replace(0, pd.NA)) * 100
        filtered_df["ratio"] = filtered_df["ratio"].fillna(0).round(2)

        # Capitalize column headers
        filtered_df.columns = filtered_df.columns.str.title()

        # Save to OUTPUT16.xlsx
        output16_file = output_folder / "OUTPUT16.xlsx"
        print(f"Saving OUTPUT16.xlsx to: {output16_file}")
        filtered_df.to_excel(output16_file, index=False)

        print(f"OUTPUT16.xlsx created successfully at {output16_file}")
        return output16_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


if __name__ == "__main__":
    # Dynamically resolve paths
    base_path = Path(__file__).resolve().parents[2]
    upload_folder = base_path / "uploads" / "REPLEN"
    output_folder = base_path / "output" / "REPLEN"

    # Run the combined script
    create_combined_output16(upload_folder, output_folder)
