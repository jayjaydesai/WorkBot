import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output17(input_file, output_file):
    """
    Update the "Decision" column in OUTPUT16.xlsx based on the "Ratio" column and specific criteria.

    Args:
        input_file (str): Path to the input Excel file (OUTPUT16.xlsx).
        output_file (str): Path to save the processed Excel file (OUTPUT17.xlsx).
    """
    try:
        # Resolve paths dynamically for compatibility
        input_file = Path(input_file).resolve()
        output_file = Path(output_file).resolve()
        print(f"Resolved input file: {input_file}")
        print(f"Resolved output file: {output_file}")

        # Validate input file existence
        if not input_file.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        # Load the Excel file into a DataFrame
        print("Loading input file...")
        df = pd.read_excel(input_file)

        # Check for missing columns
        required_columns = ["Decision", "Ratio", "Item Number"]
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Missing required column: {col}")

        # Group rows by "Item Number" where "Decision" is blank
        grouped = df[df["Decision"].isna()].groupby("Item Number")

        for item_number, group in grouped:
            ratios = group["Ratio"]
            if all(ratios < 100):  # All ratios under 100
                max_ratio_index = group["Ratio"].idxmax()
                df.loc[max_ratio_index, "Decision"] = "Good to Go"
                df.loc[group.index.difference([max_ratio_index]), "Decision"] = "Not to Use"
            elif all(ratios > 100):  # All ratios above 100
                min_ratio_index = group["Ratio"].idxmin()
                df.loc[min_ratio_index, "Decision"] = "Good to Go"
                df.loc[group.index.difference([min_ratio_index]), "Decision"] = "Not to Use"

        # Save the result to an Excel file with formatting
        print("Saving to OUTPUT17.xlsx...")
        save_with_formatting(df, output_file)
        print(f"OUTPUT17.xlsx created and saved at {output_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {e}")
        raise
    except ValueError as e:
        print(f"Value Error: {e}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def save_with_formatting(df, output_file):
    """
    Save the DataFrame to an Excel file with formatting.

    Args:
        df (pd.DataFrame): DataFrame to save.
        output_file (str): Path to save the formatted Excel file.
    """
    # Save DataFrame to Excel
    df.to_excel(output_file, index=False)

    # Load the saved workbook
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve paths for compatibility
    base_path = Path(__file__).resolve().parent.parent.parent
    input_file = base_path / "output" / "REPLEN" / "OUTPUT16.xlsx"
    output_file = base_path / "output" / "REPLEN" / "OUTPUT17.xlsx"

    # Run the function
    create_output17(input_file, output_file)
