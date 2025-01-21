import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def create_output1(upload_folder, output_folder):
    try:
        # Define required and optional files
        required_files = {
            "bulk_file": "BULK.xlsx",
            "as_file": "AS01011.xlsx",
            "coverage_file": "COVERAGE.xlsx",
        }
        optional_files = {
            "export_file": "EXPORT.xlsx",
            "threepl_file": "3PL.xlsx",
        }

        # Resolve full paths for required files
        input_files = {key: Path(upload_folder) / filename for key, filename in required_files.items()}
        optional_input_files = {key: Path(upload_folder) / filename for key, filename in optional_files.items()}

        # Validate required files exist
        missing_files = [filename for key, filename in required_files.items() if not input_files[key].exists()]
        if missing_files:
            raise FileNotFoundError(f"Missing required files: {', '.join(missing_files)}. Please upload these files to proceed.")

        # Check optional files and log warnings
        for key, file_path in optional_input_files.items():
            if not file_path.exists():
                print(f"Warning: Optional file not found: {file_path}")

        # Resolve output file path
        output_file = Path(output_folder) / "OUTPUT1.xlsx"
        os.makedirs(output_folder, exist_ok=True)  # Ensure the output folder exists

        # Load files and process data
        print("Loading BULK.xlsx...")
        bulk_df = pd.read_excel(input_files["bulk_file"])
        bulk_df.columns = bulk_df.columns.str.lower().str.strip()

        print("Loading AS01011.xlsx...")
        as_df = pd.read_excel(input_files["as_file"], usecols=["Item number", "Available physical"])
        as_df.columns = as_df.columns.str.lower().str.strip()
        as_df["available physical"] = as_df["available physical"].fillna(0)

        print("Loading COVERAGE.xlsx...")
        coverage_df = pd.read_excel(input_files["coverage_file"])
        coverage_df.columns = coverage_df.columns.str.lower().str.strip()

        # Merge BULK.xlsx with COVERAGE.xlsx on "item number"
        print("Merging BULK.xlsx and COVERAGE.xlsx...")
        merged_df = pd.merge(bulk_df, coverage_df, on="item number", how="left")

        # Merge the result with AS01011.xlsx on "item number"
        print("Merging with AS01011.xlsx...")
        final_df = pd.merge(merged_df, as_df, on="item number", how="left")

        # Save the final dataframe to OUTPUT1.xlsx
        final_df.to_excel(output_file, index=False)

        # Apply formatting to the output file
        apply_formatting(output_file)
        print(f"OUTPUT1.xlsx created and formatted at {output_file}")

        return output_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise

def apply_formatting(output_file):
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value)) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)

if __name__ == "__main__":
    # Dynamically resolve paths based on the script's environment
    base_path = Path(__file__).resolve().parent.parent.parent
    upload_folder = base_path / "uploads" / "REPLEN"
    output_folder = base_path / "output" / "REPLEN"

    # Debugging: Print resolved paths
    print(f"Resolved upload folder: {upload_folder}")
    print(f"Resolved output folder: {output_folder}")

    # Check if upload folder exists
    if not upload_folder.exists():
        print(f"Upload folder does not exist: {upload_folder}")
        raise FileNotFoundError(f"Upload folder not found: {upload_folder}")

    # List files in the upload folder
    print(f"Files in upload folder: {list(upload_folder.glob('*'))}")

    # Run the function
    create_output1(upload_folder, output_folder)


