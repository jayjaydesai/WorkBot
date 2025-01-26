import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output26(input_file, output_file):
    """
    Add a "Sr. No." column and filter columns as per the specified sequence.

    Args:
        input_file (str): Path to the input file (e.g., OUTPUT25.xlsx).
        output_file (str): Path to save the output file (e.g., OUTPUT26.xlsx).
    """
    try:
        # Resolve paths dynamically for compatibility
        input_file = Path(input_file).resolve()
        output_file = Path(output_file).resolve()
        print(f"Resolved input file: {input_file}")
        print(f"Resolved output file: {output_file}")

        # Validate input file existence
        if not input_file.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        # Load the data
        print("Loading input file...")
        df = pd.read_excel(input_file)

        # Ensure column names are consistent
        df.columns = df.columns.str.strip()

        # Add "Sr. No." column
        print("Adding 'Sr. No.' column...")
        df.insert(0, "Sr. No.", range(1, len(df) + 1))

        # Filter and rearrange columns
        print("Rearranging columns...")
        columns_to_keep = [
            "Sr. No.",
            "Item Number",
            "Location",
            "Stock Status",
            "Licence Plate",
            "Posted Quantity",
            "Final Replen Stock",
            "Rsubrange",
            "Level",
            "Missing Pallet",
            "Missing Item Number",
            "Comment",
            "Priority Status",
            "Stock",
        ]
        missing_columns = [col for col in columns_to_keep if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing columns in input file: {', '.join(missing_columns)}")
        df = df[columns_to_keep]

        # Save the updated dataframe
        print("Saving output file...")
        df.to_excel(output_file, index=False)

        # Apply formatting
        apply_formatting(output_file)
        print(f"OUTPUT26.xlsx created and formatted at {output_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {e}")
        raise
    except ValueError as e:
        print(f"Column Error: {e}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT26.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["F2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve paths for compatibility
    base_path = Path(__file__).resolve().parent.parent.parent
    input_file = base_path / "output" / "REPLEN" / "OUTPUT25.xlsx"
    output_file = base_path / "output" / "REPLEN" / "OUTPUT26.xlsx"

    # Run the function
    create_output26(input_file, output_file)
