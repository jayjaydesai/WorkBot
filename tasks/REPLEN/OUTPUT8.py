import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


def create_output8(output_folder):
    """
    Remove rows where "Ugly" column is "UGLY" unless "Stock" column is greater than 0.
    Save the result as OUTPUT8.xlsx, preserving formatting.

    Args:
        output_folder (str): Path to the folder containing OUTPUT7.xlsx.
    """
    try:
        # Resolve output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output7_file = output_folder / "OUTPUT7.xlsx"
        output8_file = output_folder / "OUTPUT8.xlsx"

        if not output7_file.exists():
            raise FileNotFoundError(f"Required file not found: {output7_file}")

        # Load OUTPUT7.xlsx
        print("Loading OUTPUT7.xlsx...")
        wb = load_workbook(output7_file)
        ws = wb.active

        # Read all rows into memory
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]  # Extract the header row

        # Ensure necessary columns exist
        if "Ugly" not in header or "Stock" not in header:
            raise ValueError("Required columns 'Ugly' or 'Stock' not found in OUTPUT7.xlsx.")

        # Get column indexes for filtering
        ugly_index = header.index("Ugly")
        stock_index = header.index("Stock")

        # Filter rows based on the condition
        print("Filtering rows based on 'Ugly' and 'Stock' conditions...")
        filtered_rows = [
            row for row in rows[1:]  # Skip header row
            if not (row[ugly_index] == "UGLY" and (row[stock_index] is None or row[stock_index] <= 0))
        ]

        # Create a new workbook for OUTPUT8.xlsx
        new_wb = Workbook()
        new_ws = new_wb.active

        # Write header and filtered rows to the new worksheet
        print("Writing filtered rows to OUTPUT8.xlsx...")
        new_ws.append(header)  # Write the header row
        for row in filtered_rows:
            new_ws.append(row)

        # Apply formatting
        print("Applying formatting to OUTPUT8.xlsx...")
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
        header_font = Font(color="FFFF00", bold=True)  # Yellow font
        alignment = Alignment(horizontal="center", vertical="center")

        for col, cell in enumerate(new_ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            col_letter = get_column_letter(col)
            new_ws.column_dimensions[col_letter].width = max(len(str(cell.value or "")) + 2, 12)

        # Apply center alignment to all data rows
        for row in new_ws.iter_rows(min_row=2, max_row=new_ws.max_row):
            for cell in row:
                cell.alignment = alignment

        # Freeze the first row and column
        new_ws.freeze_panes = "B2"

        # Save the new workbook as OUTPUT8.xlsx
        print(f"Saving OUTPUT8.xlsx to: {output8_file}")
        new_wb.save(output8_file)
        print(f"OUTPUT8.xlsx created successfully at {output8_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output8(output_folder)

