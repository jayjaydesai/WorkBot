import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output10(output_folder):
    """
    Add a "Number of Item Number" column to OUTPUT9.xlsx and save as OUTPUT10.xlsx.

    Args:
        output_folder (str): Path to the folder containing the output files.
    """
    try:
        # Resolve output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output9_file = output_folder / "OUTPUT9.xlsx"
        output10_file = output_folder / "OUTPUT10.xlsx"

        if not output9_file.exists():
            raise FileNotFoundError(f"Required file not found: {output9_file}")

        # Load OUTPUT9.xlsx
        print("Loading OUTPUT9.xlsx...")
        df = pd.read_excel(output9_file)

        # Normalize column names for consistency
        df.columns = df.columns.str.lower()

        # Validate required column
        if "item number" not in df.columns:
            raise ValueError("Required column 'Item Number' not found in OUTPUT9.xlsx.")

        # Add "Number of Item Number" column
        print("Calculating 'Number of Item Number'...")
        item_counts = df["item number"].value_counts()
        df["number of item number"] = df["item number"].map(item_counts)

        # Save the updated DataFrame to OUTPUT10.xlsx
        print(f"Saving OUTPUT10.xlsx to: {output10_file}")
        df.to_excel(output10_file, index=False)

        # Apply formatting
        print("Applying formatting to OUTPUT10.xlsx...")
        apply_formatting(output10_file)
        print(f"OUTPUT10.xlsx created and formatted successfully at {output10_file}")

        return output10_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file.

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output10(output_folder)
