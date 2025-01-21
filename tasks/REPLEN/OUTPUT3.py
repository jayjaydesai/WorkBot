import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output3(output_folder):
    """
    Process OUTPUT2.xlsx to add calculated columns and save as OUTPUT3.xlsx.

    Args:
        output_folder (str): Path to save the output file.
    """
    try:
        # Resolve paths dynamically
        output_folder = Path(output_folder).resolve()

        # Debugging: Print resolved paths
        print(f"Resolved output folder: {output_folder}")

        # Ensure folders exist
        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output2_file = output_folder / "OUTPUT2.xlsx"
        output3_file = output_folder / "OUTPUT3.xlsx"

        # Check if required file exists
        if not output2_file.exists():
            raise FileNotFoundError(f"Required file not found: {output2_file}")

        # Load OUTPUT2.xlsx
        print("Loading OUTPUT2.xlsx...")
        output2_df = pd.read_excel(output2_file)

        # Ensure case-insensitive column names
        output2_df.columns = output2_df.columns.str.lower().str.strip()

        # Add "Rmin - Available Physical" column
        print("Adding calculated columns...")
        output2_df["rmin - available physical"] = output2_df["rmin"] - output2_df["available physical"]

        # Add "Replen Status" column
        output2_df["replen status"] = output2_df["rmin - available physical"].apply(
            lambda x: "Replen Required" if x > 0 else "Replen not Required"
        )

        # Save to OUTPUT3.xlsx
        print(f"Saving OUTPUT3.xlsx to: {output3_file}")
        output2_df.to_excel(output3_file, index=False)

        # Apply formatting
        apply_formatting(output3_file)
        print(f"OUTPUT3.xlsx created and formatted successfully at {output3_file}")

        return output3_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT3.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically determine paths
    base_path = Path(__file__).resolve().parents[2]  # Go up two levels from this script's location
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output3(output_folder)
