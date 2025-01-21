import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output15(output_folder):
    """
    Process OUTPUT14.xlsx to update the "Decision" column and save as OUTPUT15.xlsx.

    Args:
        output_folder (str): Path to the folder containing the output files.
    """
    try:
        # Resolve the output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        # Check if the output folder exists
        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output14_file = output_folder / "OUTPUT14.xlsx"
        output15_file = output_folder / "OUTPUT15.xlsx"

        # Check if OUTPUT14.xlsx exists
        if not output14_file.exists():
            raise FileNotFoundError(f"{output14_file} not found.")

        # Load OUTPUT14.xlsx
        print("Loading OUTPUT14.xlsx...")
        df = pd.read_excel(output14_file)

        # Normalize column names for consistency
        df.columns = df.columns.str.lower()

        # Ensure required columns exist
        required_columns = ["decision", "item number", "pallet qty-replen stock"]
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Missing required column: {col}")

        # Process rows
        print("Processing 'Decision' column...")
        for idx, row in df.iterrows():
            if pd.isna(row["decision"]):  # Only update rows with blank "Decision"
                item_number = row["item number"]

                # Get all rows with the same "Item Number"
                same_item_rows = df[df["item number"] == item_number]

                # Check rows with "Pallet Qty-Replen Stock" == 0
                zero_pallet_rows = same_item_rows[same_item_rows["pallet qty-replen stock"] == 0]

                if not zero_pallet_rows.empty:
                    # Assign "Good to Go" for the first row where "Pallet Qty-Replen Stock" == 0
                    first_zero_idx = zero_pallet_rows.index[0]
                    df.at[first_zero_idx, "decision"] = "Good to Go"

                    # Assign "Not to Use" for rows with non-zero "Pallet Qty-Replen Stock"
                    non_zero_pallet_rows = same_item_rows[same_item_rows["pallet qty-replen stock"] != 0]
                    for non_zero_idx in non_zero_pallet_rows.index:
                        df.at[non_zero_idx, "decision"] = "Not to Use"

        # Save the updated dataframe to OUTPUT15.xlsx
        print(f"Saving OUTPUT15.xlsx to: {output15_file}")
        os.makedirs(output_folder, exist_ok=True)
        df.to_excel(output15_file, index=False)

        # Apply formatting
        print("Applying formatting to OUTPUT15.xlsx...")
        apply_formatting(output15_file)
        print(f"OUTPUT15.xlsx created and formatted successfully at {output15_file}")

        return output15_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT15.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content length
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output15(output_folder)
