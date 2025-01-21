import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output13(output_folder):
    """
    Add "Decision" column to OUTPUT12.xlsx and save as OUTPUT13.xlsx.

    Args:
        output_folder (str): Path to the folder containing the output files.
    """
    try:
        # Resolve the output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        # Check if output folder exists
        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output12_file = output_folder / "OUTPUT12.xlsx"
        output13_file = output_folder / "OUTPUT13.xlsx"

        # Check if OUTPUT12.xlsx exists
        if not output12_file.exists():
            raise FileNotFoundError(f"{output12_file} not found.")

        # Load OUTPUT12.xlsx
        print("Loading OUTPUT12.xlsx...")
        df = pd.read_excel(output12_file)

        # Normalize column names for consistency
        df.columns = df.columns.str.lower()

        # Check for the required column
        if "number of item number" not in df.columns:
            raise ValueError("Missing required column: 'Number Of Item Number'")

        # Add "Decision" column
        print("Adding 'Decision' column...")
        df["decision"] = df["number of item number"].apply(
            lambda x: "Good to Go" if x == 1 else ""
        )

        # Save to OUTPUT13.xlsx
        print(f"Saving OUTPUT13.xlsx to: {output13_file}")
        os.makedirs(output_folder, exist_ok=True)
        df.to_excel(output13_file, index=False)

        # Apply formatting
        print("Applying formatting to OUTPUT13.xlsx...")
        apply_formatting(output13_file)
        print(f"OUTPUT13.xlsx created and formatted successfully at {output13_file}")

        return output13_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT13.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output13(output_folder)

