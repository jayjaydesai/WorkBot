import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


def create_output5(output_folder):
    """
    Update "Replen Status" based on "Stock" > 0, preserve formatting from OUTPUT4.xlsx, and save as OUTPUT5.xlsx.

    Args:
        output_folder (str): Path to the folder containing OUTPUT4.xlsx.
    """
    try:
        # Resolve output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output4_file = output_folder / "OUTPUT4.xlsx"
        output5_file = output_folder / "OUTPUT5.xlsx"

        if not output4_file.exists():
            raise FileNotFoundError(f"Required file not found: {output4_file}")

        # Load OUTPUT4.xlsx
        print("Loading OUTPUT4.xlsx...")
        wb = load_workbook(output4_file)
        ws = wb.active

        # Identify relevant columns
        stock_col = None
        replen_status_col = None

        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "Stock":
                stock_col = col[0].column
            if col[0].value == "Replen Status":
                replen_status_col = col[0].column

        if stock_col is None or replen_status_col is None:
            raise ValueError("Required columns 'Stock' or 'Replen Status' not found in OUTPUT4.xlsx.")

        # Update "Replen Status" for rows where "Stock" > 0
        print("Updating 'Replen Status' column...")
        for row in range(2, ws.max_row + 1):  # Skip the header row
            stock_value = ws.cell(row=row, column=stock_col).value
            if stock_value is not None and stock_value > 0:
                ws.cell(row=row, column=replen_status_col).value = "Replen Required"

        # Format the headers
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue
        header_font = Font(color="FFFF00", bold=True)  # Yellow and bold
        center_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:  # First row (headers)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Center align all cells and adjust column widths
        print("Adjusting column widths and alignments...")
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = center_alignment
            ws.column_dimensions[col_letter].width = max_length + 2

        # Freeze first row and first column
        ws.freeze_panes = "B2"

        # Save to OUTPUT5.xlsx
        print(f"Saving OUTPUT5.xlsx to: {output5_file}")
        wb.save(output5_file)
        print(f"OUTPUT5.xlsx created successfully at {output5_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output5(output_folder)
