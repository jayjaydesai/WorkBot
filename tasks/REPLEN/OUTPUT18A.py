import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Define file paths
output17_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\OUTPUT\OUTPUT17.xlsx"
a_location_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\FINAL FILES TO SEND TO ROBERT\A_LOCATION_REPLENS.xlsx"
pallet_stacker_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\FINAL FILES TO SEND TO ROBERT\PALLET_STACKER_REPLENS.xlsx"
reach_replens_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\FINAL FILES TO SEND TO ROBERT\REACH_REPLENS.xlsx"

# Load the OUTPUT17 file
output17_df = pd.read_excel(output17_path)

# Normalize column names
output17_df.columns = output17_df.columns.str.strip().str.upper()

# Step 1: Filter rows for each LEVEL STATUS
a_location_df = output17_df[output17_df['LEVEL STATUS'] == 'A LOCATION REPLENS'].sort_values(by='LOCATION_X')
pallet_stacker_df = output17_df[output17_df['LEVEL STATUS'] == 'PALLET STACKER REPLENS'].sort_values(by='LOCATION_X')
reach_replens_df = output17_df[output17_df['LEVEL STATUS'] == 'REACH REPLENS'].sort_values(by='LOCATION_X')

# Step 2: Function to apply formatting, borders, and save
def format_and_save(df, file_path):
    # Save the DataFrame to Excel
    df.to_excel(file_path, index=False)

    # Open the workbook for formatting
    wb = load_workbook(file_path)
    ws = wb.active

    # Freeze the first row (header)
    ws.freeze_panes = "A2"

    # Apply header formatting: Dark Navy Blue background, Yellow font
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Dark Navy Blue
    header_font = Font(color="FFFF00", bold=True)  # Yellow font, bold
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add borders to all cells
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter (e.g., A, B, C)
        for cell in col:
            try:
                if cell.value:  # Only consider non-empty cells
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[col_letter].width = adjusted_width

    # Center-align all data cells
    for row in ws.iter_rows(min_row=2):  # Start from the second row
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the formatted workbook
    wb.save(file_path)
    print(f"File saved and formatted at: {file_path}")

# Step 3: Apply formatting, borders, and save all three files
format_and_save(a_location_df, a_location_path)
format_and_save(pallet_stacker_df, pallet_stacker_path)
format_and_save(reach_replens_df, reach_replens_path)
