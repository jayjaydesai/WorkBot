import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output12(output_folder):
    """
    Create OUTPUT12.xlsx by adding "Stock For Replen" column to OUTPUT11.xlsx.

    Args:
        output_folder (str): Path to the folder containing OUTPUT11.xlsx.
    """
    try:
        # Resolve output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output11_file = output_folder / "OUTPUT11.xlsx"
        output12_file = output_folder / "OUTPUT12.xlsx"

        if not output11_file.exists():
            raise FileNotFoundError(f"{output11_file} not found.")

        # Load OUTPUT11.xlsx
        print("Loading OUTPUT11.xlsx...")
        df = pd.read_excel(output11_file)

        # Normalize column names for consistency
        df.columns = df.columns.str.lower()

        # Check and handle missing columns
        required_columns = ["total posted qty", "diff2", "difference", "stock"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        # Handle missing values
        for col in required_columns:
            df[col] = df[col].fillna(0)

        # Add "Stock For Replen" column
        print("Adding 'Stock For Replen' column...")
        df["stock for replen"] = df.apply(
            lambda row: row["total posted qty"] if row["diff2"] <= 11 else row["difference"], axis=1
        )

        # Add the qty from "Stock" column to the "Stock For Replen" column
        df["stock for replen"] = df["stock for replen"] + df["stock"]

        # Save to OUTPUT12.xlsx
        print(f"Saving OUTPUT12.xlsx to: {output12_file}")
        os.makedirs(output_folder, exist_ok=True)
        df.to_excel(output12_file, index=False)

        # Apply formatting
        print("Applying formatting to OUTPUT12.xlsx...")
        apply_formatting(output12_file)
        print(f"OUTPUT12.xlsx created and formatted successfully at {output12_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file.

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze panes
    sheet.freeze_panes = sheet["B2"]

    # Format column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output12(output_folder)
