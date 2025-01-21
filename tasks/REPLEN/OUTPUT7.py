import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


def create_output7(output_folder):
    """
    Remove rows where "Stock Status" is "Available" or "Location" is "AS Replens" from OUTPUT6.xlsx,
    and save the result as OUTPUT7.xlsx with preserved formatting.

    Args:
        output_folder (str): Path to the folder containing OUTPUT6.xlsx.
    """
    try:
        # Resolve output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output6_file = output_folder / "OUTPUT6.xlsx"
        output7_file = output_folder / "OUTPUT7.xlsx"

        if not output6_file.exists():
            raise FileNotFoundError(f"Required file not found: {output6_file}")

        # Load OUTPUT6.xlsx
        print("Loading OUTPUT6.xlsx...")
        wb = load_workbook(output6_file)
        ws = wb.active

        # Read all rows into memory
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]  # Extract the header row

        # Ensure necessary columns exist
        if "Stock Status" not in header or "Location" not in header:
            raise ValueError("Required columns 'Stock Status' or 'Location' not found in OUTPUT6.xlsx.")

        # Get column indexes for filtering
        stock_status_index = header.index("Stock Status")
        location_index = header.index("Location")

        # Filter rows based on conditions
        print("Filtering rows based on 'Stock Status' and 'Location' conditions...")
        filtered_rows = [
            row for row in rows[1:]  # Skip header row
            if row[stock_status_index] != "Available" and row[location_index] != "AS Replens"
        ]

        # Create a new workbook for OUTPUT7.xlsx
        new_wb = Workbook()
        new_ws = new_wb.active

        # Write header and filtered rows to the new worksheet
        print("Writing filtered rows to OUTPUT7.xlsx...")
        new_ws.append(header)  # Write the header row
        for row in filtered_rows:
            new_ws.append(row)

        # Apply formatting
        print("Applying formatting to OUTPUT7.xlsx...")
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
        header_font = Font(color="FFFF00", bold=True)  # Yellow font
        alignment = Alignment(horizontal="center", vertical="center")

        for col, cell in enumerate(new_ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            col_letter = get_column_letter(col)
            new_ws.column_dimensions[col_letter].width = max(len(str(cell.value or "")) + 2, 12)

        # Apply center alignment to all data rows
        for row in new_ws.iter_rows(min_row=2, max_row=new_ws.max_row):
            for cell in row:
                cell.alignment = alignment

        # Freeze the first row and column
        new_ws.freeze_panes = "B2"

        # Save the new workbook as OUTPUT7.xlsx
        print(f"Saving OUTPUT7.xlsx to: {output7_file}")
        new_wb.save(output7_file)
        print(f"OUTPUT7.xlsx created successfully at {output7_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output7(output_folder)
