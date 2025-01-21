import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output19(input_file, output_file):
    """
    Update 'Decision' column for each Item Number group based on the 'Ratio' column.
    - If the nearest number to 100 is above 60, mark it as "Good to Go".
    - If the nearest number to 100 is below 60, mark the highest ratio row as "Good to Go".
    - Mark remaining rows for the same Item Number as "Not to Use".

    Args:
        input_file (str): Path to the input file (e.g., OUTPUT18.xlsx).
        output_file (str): Path to save the output file (e.g., OUTPUT19.xlsx).
    """
    try:
        # Resolve paths dynamically for compatibility
        input_file = Path(input_file).resolve()
        output_file = Path(output_file).resolve()
        print(f"Resolved input file: {input_file}")
        print(f"Resolved output file: {output_file}")

        # Validate input file existence
        if not input_file.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        # Load the data
        print("Loading input file...")
        df = pd.read_excel(input_file)

        # Ensure column names are consistent
        df.columns = df.columns.str.strip()

        # Iterate through each Item Number group
        for item_number, group in df.groupby("Item Number"):
            blank_rows = group[group["Decision"].isna()]  # Rows with blank "Decision"

            if not blank_rows.empty:
                # Find the nearest Ratio to 100
                closest_to_100_index = (blank_rows["Ratio"] - 100).abs().idxmin()
                closest_to_100_value = blank_rows.loc[closest_to_100_index, "Ratio"]

                if closest_to_100_value >= 60:
                    # If nearest to 100 is >= 60, mark it as "Good to Go"
                    df.loc[closest_to_100_index, "Decision"] = "Good to Go"
                else:
                    # If nearest to 100 is < 60, find the highest ratio row
                    highest_ratio_index = blank_rows["Ratio"].idxmax()
                    df.loc[highest_ratio_index, "Decision"] = "Good to Go"

                # Mark remaining rows for the same Item Number as "Not to Use"
                remaining_rows = blank_rows.drop(index=[closest_to_100_index] if closest_to_100_value >= 60 else [highest_ratio_index])
                df.loc[remaining_rows.index, "Decision"] = "Not to Use"

        # Save the updated dataframe
        print("Saving output file...")
        df.to_excel(output_file, index=False)

        # Apply formatting
        apply_formatting(output_file)
        print(f"OUTPUT19.xlsx created and formatted at {output_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {e}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT19.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve paths for compatibility
    base_path = Path(__file__).resolve().parent.parent.parent
    input_file = base_path / "output" / "REPLEN" / "OUTPUT18.xlsx"
    output_file = base_path / "output" / "REPLEN" / "OUTPUT19.xlsx"

    # Run the function
    create_output19(input_file, output_file)
