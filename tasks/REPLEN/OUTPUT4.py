import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font


def create_output4(output_folder):
    """
    Add a calculated "Difference" column to OUTPUT3.xlsx and save it as OUTPUT4.xlsx.
    Adjust column width, freeze panes, and apply formatting.

    Args:
        output_folder (str): Path to save the output file.
    """
    try:
        # Resolve paths dynamically
        output_folder = Path(output_folder).resolve()

        # Debugging: Print resolved paths
        print(f"Resolved output folder: {output_folder}")

        # Ensure output folder exists
        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output3_file = output_folder / "OUTPUT3.xlsx"
        output4_file = output_folder / "OUTPUT4.xlsx"

        # Check if OUTPUT3.xlsx exists
        if not output3_file.exists():
            raise FileNotFoundError(f"Required file not found: {output3_file}")

        # Load OUTPUT3.xlsx into a DataFrame
        print("Loading OUTPUT3.xlsx...")
        df = pd.read_excel(output3_file)

        # Add "Difference" column: Rcoverage - Available Physical
        print("Adding 'Difference' column...")
        df["Difference"] = df["Rcoverage"] - df["Available Physical"]

        # Save to OUTPUT4.xlsx without formatting
        print(f"Saving OUTPUT4.xlsx to: {output4_file}")
        df.to_excel(output4_file, index=False)

        # Adjust formatting in OUTPUT4.xlsx
        print("Applying formatting to OUTPUT4.xlsx...")
        adjust_formatting(output4_file)
        print(f"OUTPUT4.xlsx created and formatted successfully at {output4_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def adjust_formatting(file_path):
    """
    Adjust column widths, freeze panes, and apply formatting.

    Args:
        file_path (str): Path to the Excel file.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Style for column headers
    header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
    header_font = Font(color="FFFF00", bold=True)  # Yellow and bold

    # Center alignment
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Apply header formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Adjust column widths and align data
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = center_alignment
        ws.column_dimensions[col_letter].width = max_length + 2

    # Freeze the first row and column
    ws.freeze_panes = ws["B2"]

    # Save the workbook with adjusted formatting
    wb.save(file_path)


if __name__ == "__main__":
    # Dynamically determine paths
    base_path = Path(__file__).resolve().parents[2]  # Go up two levels from this script's location
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output4(output_folder)
