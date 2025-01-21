import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_output28(input_file, output_file):
    """
    Add formatting, highlight "Absolute Priority" in the "Priority Status" column,
    and include "Date & Time" column in OUTPUT28.xlsx.

    Args:
        input_file (str): Path to the input file (e.g., OUTPUT27.xlsx).
        output_file (str): Path to save the output file (e.g., OUTPUT28.xlsx).
    """
    try:
        # Resolve file paths dynamically
        input_file = Path(input_file).resolve()
        output_file = Path(output_file).resolve()
        print(f"Resolved input file: {input_file}")
        print(f"Resolved output file: {output_file}")

        # Validate input file existence
        if not input_file.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        # Load the data
        print("Loading input file...")
        df = pd.read_excel(input_file)

        # Add "Date & Time" column
        print("Adding 'Date & Time' column...")
        df["Date & Time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Save the updated dataframe
        print("Saving updated data...")
        df.to_excel(output_file, index=False)

        # Apply formatting and highlight cells
        apply_formatting(output_file)
        print(f"OUTPUT28.xlsx created and formatted at {output_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {e}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT28.xlsx), highlight "Absolute Priority" cells,
    and add the "Date & Time" column.

    Args:
        output_file (str): Path to the Excel file to format.
    """
    print("Applying formatting...")
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["F2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red background
    highlight_font = Font(bold=True, color="000000")  # Black bold font

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        cell.border = thin_border
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment and borders to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border

    # Highlight "Absolute Priority" in the "Priority Status" column
    priority_status_col = None
    for col_num, cell in enumerate(sheet[1], start=1):
        if cell.value == "Priority Status":
            priority_status_col = col_num
            break

    if priority_status_col:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=priority_status_col)
            if cell.value == "Absolute Priority":
                cell.fill = highlight_fill
                cell.font = highlight_font

    # Save the formatted file
    wb.save(output_file)
    print("Formatting applied and file saved.")


if __name__ == "__main__":
    # Dynamically resolve paths for compatibility
    base_path = Path(__file__).resolve().parent.parent.parent
    input_file = base_path / "output" / "REPLEN" / "OUTPUT27.xlsx"
    output_file = base_path / "output" / "REPLEN" / "OUTPUT28.xlsx"

    # Run the function
    create_output28(input_file, output_file)
