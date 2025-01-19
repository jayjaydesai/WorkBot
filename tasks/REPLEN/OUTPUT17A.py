import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import os

# Define file paths
output16_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\OUTPUT\OUTPUT16.xlsx"
output17_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\OUTPUT\OUTPUT17.xlsx"

# Load the OUTPUT16 file
output16_df = pd.read_excel(output16_path)

# Normalize column names
output16_df.columns = output16_df.columns.str.strip().str.upper()

# Step 1: Define column mappings for potential spelling differences
column_mappings = {
    "LICENSE PLATE": "LICENCE PLATE"  # Map to actual column names in the file
}

# Replace requested column names with actual column names if needed
selected_columns = [
    'ITEM NUMBER', 'LOCATION_X', 'STOCK STATUS', 'LICENSE PLATE', 'POSTED QUANTITY',
    'SUB RANGE', 'BRAND', 'LEVEL', 'FINAL STOCK FOR REPLEN', 'LEVEL STATUS',
    'SHORTAGE OF STOCK TO FULLY REPLEN'
]
adjusted_columns = [column_mappings.get(col, col) for col in selected_columns]

# Step 2: Ensure all required columns are in the file
missing_columns = [col for col in adjusted_columns if col not in output16_df.columns]
if missing_columns:
    raise KeyError(f"The following required columns are missing: {missing_columns}")

# Step 3: Create the OUTPUT17 DataFrame with only the adjusted columns
output17_df = output16_df[adjusted_columns]

# Step 4: Sort by "LEVEL STATUS" in A-to-Z order
output17_df = output17_df.sort_values(by="LEVEL STATUS", ascending=True)

# Step 5: Save the selected and sorted columns to OUTPUT17.xlsx
output17_df.to_excel(output17_path, index=False)

# Step 6: Adjust formatting: Freeze header row, add header style, center-align text
try:
    wb = load_workbook(output17_path)
    ws = wb.active

    # Freeze the first row (header)
    ws.freeze_panes = "A2"

    # Apply styles to the header row
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Navy Dark Blue background
    header_font = Font(color="FFFF00", bold=True)  # Yellow font, bold
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter (e.g., A, B, C)
        for cell in col:
            try:
                if cell.value:  # Only consider non-empty cells
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[col_letter].width = adjusted_width

    # Center-align all data cells
    for row in ws.iter_rows(min_row=2):  # Start from the second row
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the formatted OUTPUT17 file
    wb.save(output17_path)
    print(f"OUTPUT17 file saved with adjusted formatting, freeze panes, and sorting at: {output17_path}")

except Exception as e:
    print(f"An error occurred while formatting OUTPUT17: {e}")
