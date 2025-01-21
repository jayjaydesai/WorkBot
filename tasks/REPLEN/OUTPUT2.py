import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output2(upload_folder, output_folder):
    """
    Process OUTPUT1.xlsx and merge with EXPORT.xlsx to create OUTPUT2.xlsx,
    replacing blank cells in "Available Physical" with 0.

    Args:
        upload_folder (str): Path to the folder containing uploaded files.
        output_folder (str): Path to save the output file.
    """
    try:
        # Resolve paths dynamically
        upload_folder = Path(upload_folder).resolve()
        output_folder = Path(output_folder).resolve()

        # Debugging: Print resolved paths
        print(f"Resolved upload folder: {upload_folder}")
        print(f"Resolved output folder: {output_folder}")

        # Ensure folders exist
        if not upload_folder.exists():
            raise FileNotFoundError(f"Upload folder does not exist: {upload_folder}")
        output_folder.mkdir(parents=True, exist_ok=True)

        # File paths
        output1_file = output_folder / "OUTPUT1.xlsx"
        export_file = upload_folder / "EXPORT.xlsx"
        output2_file = output_folder / "OUTPUT2.xlsx"

        # Check if required files exist
        if not output1_file.exists():
            raise FileNotFoundError(f"Required file not found: {output1_file}")
        if not export_file.exists():
            print(f"Warning: EXPORT.xlsx not found. Proceeding without it.")
            export_df = pd.DataFrame(columns=["item number", "stock"])  # Create an empty DataFrame if EXPORT.xlsx is missing
        else:
            # Load EXPORT.xlsx
            print("Loading EXPORT.xlsx...")
            export_df = pd.read_excel(export_file, usecols=["Item Number", "Stock"])
            export_df.columns = export_df.columns.str.lower().str.strip()  # Normalize column names

        # Load OUTPUT1.xlsx
        print("Loading OUTPUT1.xlsx...")
        output1_df = pd.read_excel(output1_file)
        output1_df.columns = output1_df.columns.str.lower().str.strip()  # Normalize column names

        # Replace blank cells in "Available Physical" column with 0
        if "available physical" in output1_df.columns:
            output1_df["available physical"] = output1_df["available physical"].fillna(0)

        # Merge OUTPUT1.xlsx with EXPORT.xlsx on "Item Number"
        print("Merging OUTPUT1.xlsx with EXPORT.xlsx...")
        final_df = pd.merge(output1_df, export_df, on="item number", how="left")

        # Save to OUTPUT2.xlsx
        print(f"Saving OUTPUT2.xlsx to: {output2_file}")
        final_df.to_excel(output2_file, index=False)

        # Apply formatting
        apply_formatting(output2_file)
        print(f"OUTPUT2.xlsx created and formatted successfully at {output2_file}")

        return output2_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT2.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically determine paths
    base_path = Path(__file__).resolve().parents[2]  # Go up two levels from this script's location
    upload_folder = base_path / "uploads" / "REPLEN"
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output2(upload_folder, output_folder)
