import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output22(input_file, output_file):
    """
    Create OUTPUT22.xlsx:
    - Sort rows so that "Good to Go" rows under "Decision" appear at the top
      and "Not to Use" rows appear at the bottom.
    - Add a new column "Level" by extracting the suffix character from "Location".

    Args:
        input_file (str): Path to the input file (e.g., OUTPUT21.xlsx).
        output_file (str): Path to save the output file (e.g., OUTPUT22.xlsx).
    """
    try:
        # Resolve paths dynamically for compatibility
        input_file = Path(input_file).resolve()
        output_file = Path(output_file).resolve()
        print(f"Resolved input file: {input_file}")
        print(f"Resolved output file: {output_file}")

        # Validate input file existence
        if not input_file.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        # Load the data
        print("Loading input file...")
        df = pd.read_excel(input_file)

        # Ensure column names are consistent
        df.columns = df.columns.str.strip()

        # Add "Level" column by extracting the last character of the "Location" column
        if "Location" not in df.columns:
            raise ValueError("Column 'Location' not found in the input file.")
        df["Level"] = df["Location"].str[-1]

        # Handle NaN values in "Decision" and sort rows
        print("Sorting rows by 'Decision'...")
        df["Decision"] = df["Decision"].fillna("")  # Replace NaN with empty strings
        df = df.sort_values(
            by=["Decision"], 
            ascending=True, 
            key=lambda col: col.map({"Good to Go": 0, "Not to Use": 1, "": 2})
        )

        # Save the updated DataFrame to the output file
        print("Saving output file...")
        df.to_excel(output_file, index=False)

        # Apply formatting
        apply_formatting(output_file)
        print(f"OUTPUT22.xlsx created and formatted at {output_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {e}")
        raise
    except ValueError as e:
        print(f"Value Error: {e}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT22.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve paths for compatibility
    base_path = Path(__file__).resolve().parent.parent.parent
    input_file = base_path / "output" / "REPLEN" / "OUTPUT21.xlsx"
    output_file = base_path / "output" / "REPLEN" / "OUTPUT22.xlsx"

    # Run the function
    create_output22(input_file, output_file)
