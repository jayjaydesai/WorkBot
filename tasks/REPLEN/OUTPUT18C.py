import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime

# Define file paths
output17_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\OUTPUT\OUTPUT17.xlsx"
output18b_workbook_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\FINAL FILES TO SEND TO ROBERT\REPLEN_FORMAT_" + datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + ".xlsx"
a_location_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\FINAL FILES TO SEND TO ROBERT\A_LOCATION_REPLENS.xlsx"
pallet_stacker_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\FINAL FILES TO SEND TO ROBERT\PALLET_STACKER_REPLENS.xlsx"
reach_replens_path = r"C:\Users\jayde\OneDrive\AS-REPLEN-DAILY\FINAL FILES TO SEND TO ROBERT\REACH_REPLENS.xlsx"

# Load the OUTPUT17 file
output17_df = pd.read_excel(output17_path)

# Normalize column names
output17_df.columns = output17_df.columns.str.strip().str.upper()

# Step 1: Filter rows for each LEVEL STATUS
a_location_df = output17_df[output17_df['LEVEL STATUS'] == 'A LOCATION REPLENS'].sort_values(by='LOCATION_X')
pallet_stacker_df = output17_df[output17_df['LEVEL STATUS'] == 'PALLET STACKER REPLENS'].sort_values(by='LOCATION_X')
reach_replens_df = output17_df[output17_df['LEVEL STATUS'] == 'REACH REPLENS'].sort_values(by='LOCATION_X')

# Step 2: Define the required column order and additional columns
column_order = [
    'ITEM NUMBER', 'LOCATION_X', 'STOCK STATUS', 'LICENCE PLATE', 'POSTED QUANTITY',
    'FINAL STOCK FOR REPLEN', 'SUB RANGE','LEVEL', 'LEVEL STATUS'
]
renamed_columns = {
    'LOCATION_X': 'BULK LOCATION'
}

# Add SR. NO., MISSING PALLET, and MISSING ITEM columns
def prepare_dataframe(df):
    df = df[column_order]  # Select columns in the defined order
    df = df.rename(columns=renamed_columns)  # Rename columns
    df.insert(0, 'SR. NO.', range(1, len(df) + 1))  # Add SR. NO. column
    df['MISSING PALLET'] = ""  # Add empty MISSING PALLET column
    df['MISSING ITEM'] = ""  # Add empty MISSING ITEM column
    return df

a_location_df = prepare_dataframe(a_location_df)
pallet_stacker_df = prepare_dataframe(pallet_stacker_df)
reach_replens_df = prepare_dataframe(reach_replens_df)

# Step 3: Function to format each sheet or file
def format_sheet(ws):
    # Freeze the first row (header)
    ws.freeze_panes = "A2"

    # Apply header formatting: Dark Navy Blue background, Yellow font
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Dark Navy Blue
    header_font = Font(color="FFFF00", bold=True, size=10)  # Yellow font, bold
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add borders to all cells
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Adjust column widths dynamically to fit content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter (e.g., A, B, C)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding for readability

    # Center-align all data cells and reduce font size for better fit
    for row in ws.iter_rows(min_row=2):  # Start from the second row
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=10)  # Reduce font size for better fit

    # Set page layout to landscape and fit to A4 paper
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

def save_to_excel_with_formatting(df, file_path):
    df.to_excel(file_path, index=False)
    wb = load_workbook(file_path)
    ws = wb.active
    format_sheet(ws)
    wb.save(file_path)
    print(f"Formatted file saved at: {file_path}")

# Step 4: Create a single workbook with three sheets
wb = Workbook()

# Create A LOCATION REPLENS sheet
ws1 = wb.active
ws1.title = "A LOCATION REPLENS"
ws1.append(a_location_df.columns.tolist())  # Add column headers
for row in a_location_df.itertuples(index=False):
    ws1.append(row)
format_sheet(ws1)

# Create PALLET STACKER REPLENS sheet
ws2 = wb.create_sheet(title="PALLET STACKER REPLENS")
ws2.append(pallet_stacker_df.columns.tolist())  # Add column headers
for row in pallet_stacker_df.itertuples(index=False):
    ws2.append(row)
format_sheet(ws2)

# Create REACH REPLENS sheet
ws3 = wb.create_sheet(title="REACH REPLENS")
ws3.append(reach_replens_df.columns.tolist())  # Add column headers
for row in reach_replens_df.itertuples(index=False):
    ws3.append(row)
format_sheet(ws3)

# Save the workbook with three sheets
wb.save(output18b_workbook_path)
print(f"Workbook with three sheets saved at: {output18b_workbook_path}")

# Step 5: Save individual files
save_to_excel_with_formatting(a_location_df, a_location_path)
save_to_excel_with_formatting(pallet_stacker_df, pallet_stacker_path)
save_to_excel_with_formatting(reach_replens_df, reach_replens_path)
