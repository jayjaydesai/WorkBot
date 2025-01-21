import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_output11(output_folder):
    """
    Add "Diff2" and "% Required of Posted Qty" columns to OUTPUT10.xlsx and save as OUTPUT11.xlsx.

    Args:
        output_folder (str): Path to the folder containing the output files.
    """
    try:
        # Resolve output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output10_file = output_folder / "OUTPUT10.xlsx"
        output11_file = output_folder / "OUTPUT11.xlsx"

        if not output10_file.exists():
            raise FileNotFoundError(f"Required file not found: {output10_file}")

        # Load OUTPUT10.xlsx (Base file)
        print("Loading OUTPUT10.xlsx...")
        df = pd.read_excel(output10_file)

        # Normalize column names for consistency
        df.columns = df.columns.str.lower()

        # Validate required columns
        required_columns = ["total posted qty", "difference"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        # Add "Diff2" column (Total Posted Qty - Difference)
        print("Adding 'Diff2' column...")
        df["diff2"] = df["total posted qty"] - df["difference"]

        # Add "% Required of Posted Qty" column (Percentage of Difference over Total Posted Qty)
        print("Adding '% Required of Posted Qty' column...")
        df["% required of posted qty"] = (df["difference"] / df["total posted qty"]) * 100
        df["% required of posted qty"] = df["% required of posted qty"].fillna(0).round(2).astype(str) + "%"

        # Save the updated DataFrame to OUTPUT11.xlsx
        print(f"Saving OUTPUT11.xlsx to: {output11_file}")
        df.to_excel(output11_file, index=False)

        # Apply formatting
        print("Applying formatting to OUTPUT11.xlsx...")
        apply_formatting(output11_file)
        print(f"OUTPUT11.xlsx created and formatted successfully at {output11_file}")

        return output11_file

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


def apply_formatting(output_file):
    """
    Apply formatting to the Excel file (OUTPUT11.xlsx).

    Args:
        output_file (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_file)
    sheet = wb.active

    # Freeze the first row and column
    sheet.freeze_panes = sheet["B2"]

    # Styling for column headings
    heading_font = Font(bold=True, color="FFFF00")  # Yellow font
    heading_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format the first row (Column Headings)
    for col_num, cell in enumerate(sheet[1], start=1):
        cell.value = cell.value.title()  # Capitalize each word in the header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = alignment
        # Adjust column width based on content
        max_length = max((len(str(cell.value or "")) for cell in sheet[get_column_letter(col_num)]), default=0)
        sheet.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 15)

    # Apply alignment to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # Save the formatted file
    wb.save(output_file)


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output11(output_folder)
