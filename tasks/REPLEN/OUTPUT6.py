import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


def create_output6(output_folder):
    """
    Remove rows where "Replen Status" is "Replen not Required" from OUTPUT5.xlsx,
    keeping the header row, and save the result as OUTPUT6.xlsx with preserved formatting.

    Args:
        output_folder (str): Path to the folder containing OUTPUT5.xlsx.
    """
    try:
        # Resolve output folder dynamically
        output_folder = Path(output_folder).resolve()
        print(f"Resolved output folder: {output_folder}")

        if not output_folder.exists():
            raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

        # File paths
        output5_file = output_folder / "OUTPUT5.xlsx"
        output6_file = output_folder / "OUTPUT6.xlsx"

        if not output5_file.exists():
            raise FileNotFoundError(f"Required file not found: {output5_file}")

        # Load OUTPUT5.xlsx
        print("Loading OUTPUT5.xlsx...")
        wb = load_workbook(output5_file)
        ws = wb.active

        # Read all rows into memory
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]  # Extract the header row

        # Ensure "Replen Status" exists in the header
        if "Replen Status" not in header:
            raise ValueError("'Replen Status' column not found in OUTPUT5.xlsx.")

        # Filter rows where "Replen Status" is not "Replen not Required"
        replen_status_index = header.index("Replen Status")
        filtered_rows = [row for row in rows[1:] if row[replen_status_index] != "Replen not Required"]

        # Create a new workbook for OUTPUT6.xlsx
        new_wb = Workbook()
        new_ws = new_wb.active

        # Write header and filtered rows to the new worksheet
        new_ws.append(header)  # Write the header row
        for row in filtered_rows:
            new_ws.append(row)

        # Formatting
        print("Applying formatting to OUTPUT6.xlsx...")
        # Apply formatting for the header
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
        header_font = Font(color="FFFF00", bold=True)  # Yellow font
        alignment = Alignment(horizontal="center", vertical="center")

        for col, cell in enumerate(new_ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            col_letter = get_column_letter(col)
            new_ws.column_dimensions[col_letter].width = max(len(str(cell.value or "")) + 2, 12)

        # Apply center alignment to all data rows
        for row in new_ws.iter_rows(min_row=2, max_row=new_ws.max_row):
            for cell in row:
                cell.alignment = alignment

        # Freeze the first row and column
        new_ws.freeze_panes = "B2"

        # Save the new workbook as OUTPUT6.xlsx
        print(f"Saving OUTPUT6.xlsx to: {output6_file}")
        new_wb.save(output6_file)
        print(f"OUTPUT6.xlsx created successfully at {output6_file}")

    except FileNotFoundError as e:
        print(f"File Not Found Error: {str(e)}")
        raise
    except ValueError as e:
        print(f"Value Error: {str(e)}")
        raise
    except Exception as e:
        print(f"Error occurred while processing: {e}")
        raise


if __name__ == "__main__":
    # Dynamically resolve the output folder path
    base_path = Path(__file__).resolve().parent.parent.parent
    output_folder = base_path / "output" / "REPLEN"

    # Run the function
    create_output6(output_folder)
