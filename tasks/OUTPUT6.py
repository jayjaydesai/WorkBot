from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import os

def process_output5(aisle_id):
    """
    Process the OUTPUT5.xlsx file for the specified aisle by applying formatting and saving the result as OUTPUT6.

    Args:
        aisle_id (str): Aisle identifier (e.g., 'A09').

    Returns:
        str: Path to the processed OUTPUT6 file for the aisle.
    """
    try:
        # Use environment variables for folder paths
        OUTPUT_FOLDER = os.getenv("OUTPUT_FOLDER", "output")  # Default to "output"

        # File paths
        input_file = os.path.join(OUTPUT_FOLDER, f"OUTPUT5_{aisle_id}.xlsx")  # Input file for the aisle
        output_file = os.path.join(OUTPUT_FOLDER, f"OUTPUT6_{aisle_id}.xlsx")  # Output file for the aisle

        # Load the OUTPUT5 workbook
        workbook = load_workbook(input_file)
        sheet = workbook.active  # Get the active worksheet

        # Rows to be formatted
        rows_to_format = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30]

        # Define the font, alignment, and fill styles
        custom_font = Font(size=10, bold=True)  # Font size 10, bold
        custom_alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray fill

        # Apply styles to the specified rows, excluding column A
        for row in rows_to_format:
            # Set row height
            sheet.row_dimensions[row].height = 11
            for col in range(2, sheet.max_column + 1):  # Start from column B
                cell = sheet.cell(row=row, column=col)
                cell.font = custom_font
                cell.alignment = custom_alignment
                cell.fill = light_gray_fill

        # Save the updated workbook as OUTPUT6
        workbook.save(output_file)
        print(f"Processed file for {aisle_id} saved as {output_file}")

        return output_file

    except Exception as e:
        print(f"Error processing aisle {aisle_id}: {e}")
        raise Exception(f"Failed to process OUTPUT6 for aisle {aisle_id}: {e}")
