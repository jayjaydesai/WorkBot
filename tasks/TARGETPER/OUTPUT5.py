import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Define dynamic paths
base_path = os.getenv("BASE_UPLOAD_PATH", os.path.join("C:", os.sep, "Users", "jaydi", "OneDrive - Comline", "CAPLOCATION", "Deployment", "bulk_report_webapp", "output", "TARGETPER"))
output4_file = os.path.join(base_path, "OUTPUT4.xlsx")
output5_file = os.path.join(base_path, "TARGET PERFORMANCE.xlsx")

# Ensure the input file exists
if not os.path.exists(output4_file):
    raise FileNotFoundError(f"INPUT FILE NOT FOUND: {output4_file}")

# Load the Excel file
try:
    wb = load_workbook(output4_file)
    print(f"SUCCESS: Loaded workbook '{output4_file}'")
except Exception as e:
    raise Exception(f"ERROR: Could not load workbook. Details: {e}")

# Check for the presence of 'main' and 'summary' sheets
if 'main' not in wb.sheetnames or 'summary' not in wb.sheetnames:
    print(f"AVAILABLE SHEETS: {wb.sheetnames}")
    raise KeyError("ERROR: The required sheets 'main' or 'summary' are missing in OUTPUT4.xlsx.")

main_sheet = wb['main']
summary_sheet = wb['summary']

# MAIN Sheet Modifications
print("PROCESSING: MAIN Sheet")
df_main = pd.DataFrame(main_sheet.values)

# Remove 'standardcode' column and add 'SR.NO.'
df_main.columns = df_main.iloc[0]  # Set the first row as headers
df_main = df_main[1:]  # Remove the first row
df_main.insert(0, 'SR.NO.', range(1, len(df_main) + 1))  # Add serial numbers
df_main.drop(columns=['standardcode'], inplace=True, errors='ignore')
df_main.rename(columns={
    'partnumber': 'PART NUMBER',
    'customercode': 'CUSTOMER CODE',
    'customername': 'CUSTOMER NAME',
    'targetdate': 'TARGET DATE',
    'status': 'STATUS',
    '2023-customer-sales': 'SALES(2023)',
    '2024-customer-sales': 'SALES(2024)',
    'number-of-customer-target': 'TARGET COUNT BY CUSTOMER',
    'number-of-part-number-target': 'TARGET COUNT BY PART',
    'sales-after-target': 'SALES AFTER TARGET'
}, inplace=True)
df_main.drop(columns=['total-sales-after-first-target'], inplace=True, errors='ignore')

# Format TARGET DATE to dd-mm-yyyy
df_main['TARGET DATE'] = pd.to_datetime(df_main['TARGET DATE'], format='%d-%m-%Y', errors='coerce').dt.strftime('%d-%m-%Y')

# Ensure any invalid dates are handled (if required)
if df_main['TARGET DATE'].isnull().any():
    print("WARNING: Some TARGET DATE values were invalid and set to NaT.")


# Write back headers and data to the MAIN sheet
main_sheet.delete_rows(1, main_sheet.max_row)  # Clear the sheet
main_sheet.append(df_main.columns.tolist())  # Add headers
for row in df_main.values:
    main_sheet.append(row.tolist())

# Apply formatting to the MAIN sheet
header_font = Font(bold=True, color="FFFF00")
header_fill = PatternFill("solid", fgColor="1F4E78")
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for cell in main_sheet[1]:  # Format headers
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# Set column widths, apply center alignment, and borders to all cells
for col in main_sheet.columns:
    column_width = 15
    col_letter = get_column_letter(col[0].column)
    main_sheet.column_dimensions[col_letter].width = column_width
    for cell in col:
        cell.alignment = center_alignment
        cell.border = thin_border

# Freeze panes at cell G2
main_sheet.freeze_panes = 'G2'
print("SUCCESS: Processed MAIN Sheet")

# SUMMARY Sheet Modifications
print("PROCESSING: SUMMARY Sheet")
df_summary = pd.DataFrame(summary_sheet.values)

# Rename first column and add 'SR.NO.'
df_summary.columns = df_summary.iloc[0]
df_summary = df_summary[1:]
df_summary.insert(0, 'SR.NO.', range(1, len(df_summary) + 1))  # Add serial numbers
df_summary.rename(columns={'customercode': 'CUSTOMER CODE'}, inplace=True)
df_summary.columns = [col.upper().replace("_", " ") for col in df_summary.columns]  # Convert all headers to uppercase and replace "_" with " "

# Rename specific columns
df_summary.rename(columns={'TOTAL BEST OFFERED': 'BEST OFFERED'}, inplace=True)

# Write back headers and data to the SUMMARY sheet
summary_sheet.delete_rows(1, summary_sheet.max_row)  # Clear the sheet
summary_sheet.append(df_summary.columns.tolist())  # Add headers
for row in df_summary.values:
    summary_sheet.append(row.tolist())

# Apply formatting to the SUMMARY sheet
for cell in summary_sheet[1]:  # Format headers
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# Set specific column widths
column_widths = {
    "A": 7,  # SR.NO.
    "B": 17,  # CUSTOMER CODE
    "C": 7,  # Q1
    "D": 7,  # Q2
    "E": 7,  # Q3
    "F": 7,  # Q4
    "G": 10,  # TOTAL
    "H": 17,  # TARGET ACCEPTED
    "I": 17,  # BEST OFFERED
    "J": 17   # NO DISCOUNT
}

# Apply column widths, alignment, and borders
for col_letter, width in column_widths.items():
    summary_sheet.column_dimensions[col_letter].width = width  # Set column width
    for cell in summary_sheet[col_letter]:
        cell.alignment = center_alignment  # Center alignment
        cell.border = thin_border  # Thin border

# Freeze panes at cell K2
summary_sheet.freeze_panes = 'K2'
print("SUCCESS: Processed SUMMARY Sheet")

# Save the updated Excel file
wb.save(output5_file)
print(f"TARGET PERFORMANCE.xlsx has been successfully generated at: {output5_file}")


