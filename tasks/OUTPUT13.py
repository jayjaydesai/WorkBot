import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Function to get the BULK.xlsx file directly
def get_bulk_file(upload_folder):
    bulk_file = Path(upload_folder) / "BULK.xlsx"
    print(f"Checking for BULK.xlsx at: {bulk_file}")
    if not bulk_file.exists():
        raise FileNotFoundError(f"BULK.xlsx not found in the uploads folder: {upload_folder}")
    return bulk_file

def process_files():
    # Set base_path dynamically to point to the `bulk_report_webapp` directory
    base_path = Path(__file__).resolve().parent.parent

    # Define folder structure dynamically
    locations_folder = base_path / "locations"
    uploads_folder = base_path / "uploads"
    output_folder = base_path / "output"

    print(f"Uploads folder: {uploads_folder}")
    print(f"Output folder: {output_folder}")

    # Ensure output folder exists
    output_folder.mkdir(parents=True, exist_ok=True)

    # Paths for input files
    a16_path = locations_folder / "A16.xlsx"
    bulk_path = get_bulk_file(uploads_folder)

    # Load Excel files
    a16_df = pd.read_excel(a16_path)
    bulk_df = pd.read_excel(bulk_path)

    # Merge based on the "Location" column, keeping only the Licence plate from BULK
    merged_df = pd.merge(
        a16_df.drop(columns=['Licence plate']),
        bulk_df[['Location', 'Licence plate']],
        on='Location',
        how='left'
    )

    # Filter rows where Licence plate is empty
    filtered_df = merged_df[merged_df['Licence plate'].isna()].copy()

    # Reset Sr.No. column to maintain sequence
    filtered_df['Sr.No.'] = range(1, len(filtered_df) + 1)

    # Add a "Date and Time" column
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filtered_df['Date and Time'] = current_datetime

    # Save filtered DataFrame to Excel
    output_file = output_folder / "EMPTYLOCATION.xlsx"
    filtered_df.to_excel(output_file, index=False)

    # Add formatting and summary sheet
    format_sheets(output_file, filtered_df, current_datetime)

    print(f"Formatted output saved to: {output_file}")

def format_sheets(file_path, filtered_df, current_datetime):
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Load workbook
    wb = load_workbook(file_path)

    # Define common styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFF00")
    border_style = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: "EMPTY LOCATION"
    sheet1 = wb.active
    sheet1.title = "EMPTY LOCATION"

    # Format header row in "EMPTY LOCATION"
    for cell in sheet1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style

    # Add borders and alignment to all data cells in "EMPTY LOCATION"
    for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
        for cell in row:
            cell.border = border_style
            cell.alignment = alignment_center

    # Adjust column widths for "EMPTY LOCATION"
    column_widths_sheet1 = {
        "A": 10,  # Sr.No.
        "B": 15,  # Aisle No
        "C": 10,  # Level
        "D": 25,  # Location
        "E": 20,  # Licence plate
        "F": 20,  # Date and Time
    }
    for col, width in column_widths_sheet1.items():
        sheet1.column_dimensions[col].width = width

    # Freeze panes at F2 in "EMPTY LOCATION"
    sheet1.freeze_panes = "F2"

    # Create a new summary sheet
    summary_ws = wb.create_sheet(title="Summary")

    # Add heading: "SUMMARY OF EMPTY LOCATIONS"
    summary_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    heading_cell = summary_ws.cell(row=1, column=1)
    heading_cell.value = "SUMMARY OF EMPTY LOCATIONS"
    heading_cell.fill = header_fill
    heading_cell.font = header_font
    heading_cell.alignment = alignment_center
    heading_cell.border = border_style

    # Add "GRAND TOTAL" row with the total number of empty locations and Date/Time
    grand_total = filtered_df.shape[0]
    summary_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    grand_total_cell = summary_ws.cell(row=2, column=1)
    grand_total_cell.value = f"GRAND TOTAL OF EMPTY LOCATION - {grand_total} (Date: {current_datetime})"
    grand_total_cell.fill = header_fill
    grand_total_cell.font = header_font
    grand_total_cell.alignment = alignment_center
    grand_total_cell.border = border_style

    # Apply borders to the first two rows
    for row in summary_ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=9):
        for cell in row:
            cell.border = border_style

    # Add table headers (Row 3)
    summary_ws.append([
        "AISLE NO", "LEVEL", "EMPTY COUNT", "", "AISLE NO.", "TOTAL EMPTY", "", "LEVEL", "TOTAL EMPTY"
    ])
    for cell in summary_ws[3]:
        cell.alignment = alignment_center
        cell.font = header_font
        # Add background color and border unless column is D or G
        if cell.column_letter not in ["D", "G"]:
            cell.fill = header_fill
            cell.border = border_style

    # Calculate summaries
    aisle_level_summary = filtered_df.groupby(["Aisle No", "Level"]).size().reset_index(name="Empty Count")
    total_by_aisle = aisle_level_summary.groupby("Aisle No")["Empty Count"].sum().reset_index()
    total_by_level = aisle_level_summary.groupby("Level")["Empty Count"].sum().reset_index()

    # Populate data for aisle-level details (Table 1)
    row_num = 4
    for _, row in aisle_level_summary.iterrows():
        summary_ws.cell(row=row_num, column=1, value=row["Aisle No"])
        summary_ws.cell(row=row_num, column=2, value=row["Level"])
        summary_ws.cell(row=row_num, column=3, value=row["Empty Count"])
        row_num += 1

    # Populate data for aisle totals (Table 2)
    row_num = 4
    for _, row in total_by_aisle.iterrows():
        summary_ws.cell(row=row_num, column=5, value=row["Aisle No"])
        summary_ws.cell(row=row_num, column=6, value=row["Empty Count"])
        row_num += 1

    # Populate data for level totals (Table 3)
    row_num = 4
    for _, row in total_by_level.iterrows():
        summary_ws.cell(row=row_num, column=8, value=row["Level"])
        summary_ws.cell(row=row_num, column=9, value=row["Empty Count"])
        row_num += 1

    # Apply borders and alignment to all data rows, excluding empty cells in specific columns
    for row in summary_ws.iter_rows(min_row=4, max_row=summary_ws.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = alignment_center
            # Remove borders for empty cells in columns E, F, H, and I
            if cell.column_letter in ["E", "F", "H", "I"] and cell.value is None:
                continue
            if cell.column_letter not in ["D", "G"]:
                cell.border = border_style

    # Adjust column widths for Summary Sheet
    column_widths_summary = {
        "A": 15,
        "B": 15,
        "C": 15,
        "E": 20,
        "F": 15,
        "H": 15,
        "I": 15,
    }
    for col, width in column_widths_summary.items():
        summary_ws.column_dimensions[col].width = width

    # Freeze panes at I4 in Summary Sheet
    summary_ws.freeze_panes = "I4"

    # Save the updated workbook
    wb.save(file_path)

# Example of using the script dynamically
if __name__ == "__main__":
    process_files()



