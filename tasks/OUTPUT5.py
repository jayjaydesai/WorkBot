from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os

def process_output4(aisle_id):
    """
    Process the OUTPUT4.xlsx file for the specified aisle by applying formatting and saving the result as OUTPUT5.

    Args:
        aisle_id (str): Aisle identifier (e.g., 'A09').

    Returns:
        str: Path to the processed OUTPUT5 file for the aisle.
    """
    try:
        # Use environment variables for folder paths
        output_folder = os.getenv("OUTPUT_FOLDER", "output")  # Default to "output" if not set

        # File paths
        input_file = os.path.join(output_folder, f"OUTPUT4_{aisle_id}.xlsx")  # Input file for the aisle
        output_file = os.path.join(output_folder, f"OUTPUT5_{aisle_id}.xlsx")  # Output file for the aisle

        # Load the OUTPUT4 workbook
        workbook = load_workbook(input_file)
        sheet = workbook.active  # Get the active worksheet

        # Rows to format
        rows_to_format = [3, 5, 7, 9, 11, 13, 15, 19, 21, 23, 25, 27, 29, 31]

        # Define styles
        default_font = Font(size=8, bold=False)  # Font size 8, non-bold
        default_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)  # Left-top aligned with wrapping

        # Apply styles to the specified rows and columns
        for row in rows_to_format:
            # Set row height to 35
            sheet.row_dimensions[row].height = 35
            for col in range(1, sheet.max_column + 1):  # Iterate through all columns
                cell = sheet.cell(row=row, column=col)
                cell.font = default_font  # Set font size 8, non-bold
                cell.alignment = default_alignment  # Align text to left and top

        # Save the updated workbook as OUTPUT5
        workbook.save(output_file)
        print(f"Processed file for {aisle_id} saved as {output_file}")

        return output_file

    except Exception as e:
        print(f"Error processing aisle {aisle_id}: {e}")
        raise
