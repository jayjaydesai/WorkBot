import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

# ✅ Define dynamic paths for local and Azure compatibility
BASE_DIR = os.getenv("BASE_DIR")
if not BASE_DIR:
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

OUTPUT_PATH = os.path.join(BASE_DIR, "output", "GREPLEN")

# ✅ Ensure required directories exist
os.makedirs(OUTPUT_PATH, exist_ok=True)

# ✅ Define file paths
input_file = os.path.join(OUTPUT_PATH, "OUTPUT25.csv")
output_file = os.path.join(OUTPUT_PATH, "WORKING.xlsx")

# ✅ Ensure the input file exists
if not os.path.exists(input_file):
    raise FileNotFoundError(f"❌ ERROR: The file '{input_file}' does not exist. Please run OUTPUT25.py first.")

# ✅ Load the source CSV
df = pd.read_csv(input_file)

# ✅ Save to Excel to retain the data structure for formatting
df.to_excel(output_file, index=False, sheet_name="Sheet1")

# ✅ Load the workbook for styling
wb = load_workbook(output_file)
sheet = wb.active

# ✅ Define border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# ✅ Styling: Header formatting
header_font = Font(size=8, bold=True, color="FFFF00")  # Yellow font
header_fill = PatternFill("solid", fgColor="1F4E78")  # Dark Blue background
center_alignment = Alignment(horizontal="center", vertical="center")

for cell in sheet[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# ✅ Set column widths
for col in sheet.columns:
    col_letter = get_column_letter(col[0].column)
    sheet.column_dimensions[col_letter].width = 10  # Default width for all columns
sheet.column_dimensions["E"].width = 28  # COMPANY NAME column
sheet.column_dimensions["U"].width = 38  # NOTE column
sheet.column_dimensions["V"].width = 48  # REASON column

# ✅ Freeze the header row
sheet.freeze_panes = "G2"

# ✅ Apply formatting to all data cells (excluding headers)
data_font = Font(size=10)  # Set all data text to font size 10
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.font = data_font  # Set font size
        cell.alignment = center_alignment  # Center align text
        cell.border = thin_border  # Apply border to each cell

# ✅ Center align all data cells
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.alignment = center_alignment

# ✅ Conditional formatting for "NOTE" column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=21, max_col=21):  # NOTE column (U, 21st column)
    cell = row[0]
    if cell.value == "Full Allocation" or cell.value == "Best Possible Allocated":
        cell.fill = PatternFill("solid", fgColor="006400")  # Dark Green background
        cell.font = Font(color="FFFF00")  # Yellow font
    elif cell.value == "Please check as ETA is closer so not allocated":
        cell.fill = PatternFill("solid", fgColor="FFA07A")  # Light Orange background

# ✅ Remove rows where "PART NUMBER" column is blank
df = df[df["PART NUMBER"].notna() & (df["PART NUMBER"].astype(str).str.strip() != "")]

# ✅ Save the formatted workbook
wb.save(output_file)
print(f"SUCCESS: WORKING.xlsx has been generated with all formatting at {output_file}")
