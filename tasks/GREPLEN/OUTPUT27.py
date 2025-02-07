import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ✅ Define dynamic paths for local and Azure compatibility
BASE_DIR = os.getenv("BASE_DIR")
if not BASE_DIR:
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

OUTPUT_PATH = os.path.join(BASE_DIR, "output", "GREPLEN")

# ✅ Ensure required directories exist
os.makedirs(OUTPUT_PATH, exist_ok=True)

# ✅ Define file paths
input_file = os.path.join(OUTPUT_PATH, "WORKING.xlsx")
output_file = os.path.join(OUTPUT_PATH, "FINAL_GR_BO_EXPORT.xlsx")

# ✅ Ensure input file exists
if not os.path.exists(input_file):
    raise FileNotFoundError(f"❌ ERROR: The file '{input_file}' does not exist. Please run the previous script first.")

# ✅ Load workbook
wb = load_workbook(input_file)
sheet = wb.active

# ✅ Load data into pandas DataFrame (preserving headers)
df = pd.DataFrame(sheet.values)
df.columns = df.iloc[0]  # Set first row as column headers
df = df[1:]  # Remove header row from data

# ✅ Remove rows where "RELEASE" column is zero (ensure type consistency)
if "RELEASE" in df.columns:
    df = df[df["RELEASE"].astype(str).str.strip() != "0"]

# ✅ Write back data to the same workbook, preserving formatting
wb.remove(sheet)  # Remove existing sheet
ws = wb.create_sheet("FINAL_GR_BO_EXPORT")  # Create new sheet

# ✅ Write headers
ws.append(df.columns.tolist())

# ✅ Write rows with formatting preserved
for row in df.itertuples(index=False):
    ws.append(row)

# ✅ Apply formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

header_font = Font(size=8, bold=True, color="FFFF00")  # Yellow font
header_fill = PatternFill("solid", fgColor="1F4E78")  # Dark Blue background
center_alignment = Alignment(horizontal="center", vertical="center")

# ✅ Apply header formatting
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# ✅ Set default column width
for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = 10  # Default width

# ✅ Set specific column widths
ws.column_dimensions["E"].width = 28  # COMPANY NAME column
ws.column_dimensions["U"].width = 38  # NOTE column
ws.column_dimensions["V"].width = 48  # REASON column

# ✅ Freeze header row
ws.freeze_panes = "G2"

# ✅ Apply formatting to all data cells (excluding headers)
data_font = Font(size=10)  # Set all data text to font size 10
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = data_font  # Set font size
        cell.alignment = center_alignment  # Center align text
        cell.border = thin_border  # Apply border to each cell

# ✅ Conditional formatting for "NOTE" column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=21, max_col=21):  # NOTE column is column U (21st column)
    cell = row[0]
    if cell.value == "Full Allocation" or cell.value == "Best Possible Allocated":
        cell.fill = PatternFill("solid", fgColor="006400")  # Dark Green background
        cell.font = Font(color="FFFF00")  # Yellow font
    elif cell.value == "Please check as ETA is closer so not allocated":
        cell.fill = PatternFill("solid", fgColor="FFA07A")  # Light Orange background

# ✅ Save final workbook
wb.save(output_file)
print(f"SUCCESS: FINAL_GR_BO_EXPORT.xlsx has been generated at {output_file}")
