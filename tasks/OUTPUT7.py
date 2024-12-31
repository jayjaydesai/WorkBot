from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import os

def process_output6(aisle_id):
    """
    Process the OUTPUT6.xlsx file for the specified aisle by applying enhanced walking path styling 
    and saving the result as OUTPUT7.

    Args:
        aisle_id (str): Aisle identifier (e.g., 'A09').

    Returns:
        str: Path to the processed OUTPUT7 file for the aisle.
    """
    try:
        # Use environment variables for folder paths
        OUTPUT_FOLDER = os.getenv("OUTPUT_FOLDER", "output")  # Default to "output"

        # File paths
        input_file = os.path.join(OUTPUT_FOLDER, f"OUTPUT6_{aisle_id}.xlsx")  # Input file for the aisle
        output_file = os.path.join(OUTPUT_FOLDER, f"OUTPUT7_{aisle_id}.xlsx")  # Output file for the aisle

        # Load the OUTPUT6 workbook
        workbook = load_workbook(input_file)
        sheet = workbook.active  # Get the active worksheet

        # Define styles
        fill_light_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        bold_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )
        arrow_font = Font(size=12, bold=True, color="1F4E78")  # Bold arrow font
        central_alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

        # Apply styles to row 16, columns A to AD
        sheet.row_dimensions[16].height = 25  # Adjust row height
        for col in range(1, 31):  # Columns A to AD
            cell = sheet.cell(row=16, column=col)
            
            # Alternating colors for the path
            if col % 2 == 0:
                cell.fill = fill_light_gray
            else:
                cell.fill = fill_white
            
            # Add directional arrow and style
            cell.value = "→"
            cell.font = arrow_font
            cell.alignment = central_alignment
            cell.border = bold_border

        # Save the updated workbook as OUTPUT7
        workbook.save(output_file)
        print(f"Processed file for {aisle_id} saved as {output_file}")

        return output_file

    except Exception as e:
        print(f"Error processing aisle {aisle_id}: {e}")
        raise Exception(f"Failed to process OUTPUT7 for aisle {aisle_id}: {e}")
